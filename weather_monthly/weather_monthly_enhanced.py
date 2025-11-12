#!/usr/bin/env python3
"""
Enhanced Weather Data Monthly Extractor for EEC Rain Gauge System
รองรับการดึงข้อมูล:
- อุณหภูมิรายวัน (Daily Temperature)
- ความชื้นรายวัน (Daily Humidity)
- ปริมาณน้ำฝน 24 ชั่วโมง (24-hour Rainfall)
- Export เป็น Excel และ CSV พร้อมกราฟ
"""

import re
import json
import csv
import html as _html
from datetime import datetime, timedelta, timezone
import requests
from bs4 import BeautifulSoup
import time
import random
import pandas as pd
import numpy as np
from typing import Optional, List, Dict, Any
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList

# URLs Configuration
LOGIN_URL = "http://eecmobile1.fortiddns.com/eec/Login.aspx"
MAP_URL = "http://eecmobile1.fortiddns.com/eec/Raingauge_Monitor_Map.aspx"
SUMMARY_URL = "http://eecmobile1.fortiddns.com/eec/Raingauge_Summary_Station.aspx"
ALL_LATEST_URL = "http://eecmobile1.fortiddns.com/eec/Raingauge_All_Lastest.aspx"

# Thai month names
THAI_MONTHS = {
    1: "มกราคม", 2: "กุมภาพันธ์", 3: "มีนาคม", 4: "เมษายน",
    5: "พฤษภาคม", 6: "มิถุนายน", 7: "กรกฎาคม", 8: "สิงหาคม",
    9: "กันยายน", 10: "ตุลาคม", 11: "พฤศจิกายน", 12: "ธันวาคม"
}

# ---------------- Helper Functions from original code ----------------
def request_with_retry(session: requests.Session, method: str, url: str,
                       *, max_attempts: int = 7,
                       base_sleep: float = 0.8,
                       timeout: float = 60,
                       retry_http_status=(500, 502, 503, 504),
                       debug: bool = False,
                       **kwargs) -> requests.Response:
    """Retry on server errors, timeouts, and transient network faults."""
    attempt = 0
    last_err = None
    while attempt < max_attempts:
        attempt += 1
        try:
            resp = session.request(method, url, timeout=timeout, **kwargs)
            if resp.status_code in retry_http_status:
                if debug:
                    print(f"[DEBUG] {url} -> HTTP {resp.status_code} on attempt {attempt}/{max_attempts}")
                raise requests.exceptions.HTTPError(f"{resp.status_code} Server Error", response=resp)
            return resp
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError,
                requests.exceptions.HTTPError) as e:
            last_err = e
            if attempt >= max_attempts:
                break
            sleep_s = base_sleep * (2 ** (attempt - 1)) + random.random() * base_sleep
            if debug:
                print(f"[DEBUG] Retry {attempt}/{max_attempts} after error: {e}. Sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)
    if last_err:
        raise last_err
    raise RuntimeError("request_with_retry failed without exception")

def _inputs(html):
    """Parse form inputs from HTML"""
    soup = BeautifulSoup(html, "html.parser")
    return {i.get("name"): i.get("value", "") for i in soup.find_all("input") if i.get("name")}

def login(user: str, password: str, debug: bool = False) -> requests.Session:
    """Login to EEC system"""
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
    
    if debug:
        print(f"[DEBUG] GET {LOGIN_URL}")
    
    r = request_with_retry(s, 'GET', LOGIN_URL, debug=debug)
    r.raise_for_status()
    
    data = _inputs(r.text)
    data.update({"tb_user": user, "tb_password": password})
    
    if debug:
        print(f"[DEBUG] POST {LOGIN_URL}")
    
    r2 = request_with_retry(s, 'POST', LOGIN_URL, data=data, allow_redirects=True, debug=debug)
    
    if "Default.aspx" in r2.url or "logout" in r2.text.lower():
        if debug:
            print("[DEBUG] Login OK")
        return s
    
    raise RuntimeError("Login failed")

# ---------------- Station Functions ----------------
def fetch_all_stations(session: requests.Session, debug: bool = False) -> Dict[str, Dict]:
    """Fetch all stations with their current status"""
    stations_dict = {}
    
    try:
        if debug:
            print(f"[DEBUG] Fetching all stations from {ALL_LATEST_URL}")
        
        response = request_with_retry(session, 'GET', ALL_LATEST_URL, debug=debug)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Find GridView table
        table = soup.find('table', {'id': 'GridView1'}) or \
                soup.find('table', {'class': 'GridView'})
        
        if not table:
            if debug:
                print("[DEBUG] No GridView table found")
            return stations_dict
        
        rows = table.find_all('tr')
        
        for row in rows[1:]:  # Skip header
            cells = row.find_all('td')
            if len(cells) >= 8:
                station_code = cells[0].get_text(strip=True)
                station_name = cells[1].get_text(strip=True)
                status = cells[2].get_text(strip=True)
                rain = cells[3].get_text(strip=True)
                temp = cells[4].get_text(strip=True)
                humidity = cells[5].get_text(strip=True)
                battery = cells[6].get_text(strip=True)
                last_update = cells[7].get_text(strip=True)
                
                stations_dict[station_code] = {
                    'code': station_code,
                    'name': station_name,
                    'status': status,
                    'last_rain': rain,
                    'last_temp': temp,
                    'last_humidity': humidity,
                    'battery': battery,
                    'last_update': last_update
                }
        
        if debug:
            print(f"[DEBUG] Found {len(stations_dict)} stations")
        
    except Exception as e:
        if debug:
            print(f"[DEBUG] Error fetching stations: {e}")
    
    return stations_dict

# ---------------- Monthly Data Fetching ----------------
def fetch_station_daily_data(session: requests.Session, station_id: str,
                            start_date: datetime, end_date: datetime,
                            debug: bool = False) -> List[Dict]:
    """Fetch daily weather data for a station between dates"""
    
    daily_data = []
    current_date = start_date
    
    print(f"\n📊 Fetching data for station {station_id}")
    print(f"   Period: {start_date.date()} to {end_date.date()}")
    
    while current_date <= end_date:
        try:
            # Format date for URL parameter
            date_str = current_date.strftime('%d/%m/%Y')
            
            # Try to fetch from Summary Station page with specific date
            url = f"{SUMMARY_URL}?id={station_id}&date={date_str}"
            
            if debug:
                print(f"[DEBUG] Fetching {url}")
            
            response = request_with_retry(session, 'GET', url, timeout=30, debug=debug)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Initialize data for this day
                day_data = {
                    'date': current_date.strftime('%Y-%m-%d'),
                    'temperature_c': None,
                    'humidity_pct': None,
                    'rain_24h_mm': None,
                    'battery_v': None,
                    'solar_v': None
                }
                
                # Parse data from page
                text_content = soup.get_text()
                
                # Temperature
                temp_match = re.search(r'Temperature\s*:\s*([+-]?\d+(?:\.\d+)?)\s*°?C?', text_content, re.I)
                if temp_match:
                    day_data['temperature_c'] = float(temp_match.group(1))
                
                # Humidity
                humidity_match = re.search(r'Humidity\s*:\s*(\d+(?:\.\d+)?)\s*%?', text_content, re.I)
                if humidity_match:
                    day_data['humidity_pct'] = float(humidity_match.group(1))
                
                # Rain
                rain_match = re.search(r'Rain(?:\s*24\s*hr)?\s*:\s*(\d+(?:\.\d+)?)\s*mm', text_content, re.I)
                if rain_match:
                    day_data['rain_24h_mm'] = float(rain_match.group(1))
                
                # Battery
                battery_match = re.search(r'Battery\s*:\s*(\d+(?:\.\d+)?)\s*V?', text_content, re.I)
                if battery_match:
                    day_data['battery_v'] = float(battery_match.group(1))
                
                # Solar
                solar_match = re.search(r'Solar\s*(?:Panel)?\s*:\s*(\d+(?:\.\d+)?)\s*V?', text_content, re.I)
                if solar_match:
                    day_data['solar_v'] = float(solar_match.group(1))
                
                daily_data.append(day_data)
                
                # Show progress
                if not debug:
                    print(f"   ✓ {current_date.strftime('%Y-%m-%d')}: "
                          f"T={day_data['temperature_c']}°C, "
                          f"H={day_data['humidity_pct']}%, "
                          f"R={day_data['rain_24h_mm']}mm")
            
            else:
                # Add empty data for this day
                daily_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'temperature_c': None,
                    'humidity_pct': None,
                    'rain_24h_mm': None,
                    'battery_v': None,
                    'solar_v': None
                })
            
            # Small delay to avoid overwhelming server
            time.sleep(0.3)
            
        except Exception as e:
            if debug:
                print(f"[DEBUG] Error fetching data for {current_date.date()}: {e}")
            
            # Add empty data for failed day
            daily_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'temperature_c': None,
                'humidity_pct': None,
                'rain_24h_mm': None,
                'battery_v': None,
                'solar_v': None
            })
        
        current_date += timedelta(days=1)
    
    return daily_data

# ---------------- Enhanced Excel Export with Charts ----------------
def export_to_excel_with_charts(data: List[Dict], station_info: Dict,
                               year: int, month: int,
                               filename: str = None) -> str:
    """Export data to Excel with formatting and charts"""
    
    if filename is None:
        filename = f"weather_{station_info['code']}_{year}_{month:02d}.xlsx"
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Ensure date column is datetime
    df['date'] = pd.to_datetime(df['date'])
    df = df.sort_values('date')
    
    # Add day of month column
    df['day'] = df['date'].dt.day
    
    # Calculate statistics
    stats = {
        'avg_temp': df['temperature_c'].mean(),
        'min_temp': df['temperature_c'].min(),
        'max_temp': df['temperature_c'].max(),
        'avg_humidity': df['humidity_pct'].mean(),
        'min_humidity': df['humidity_pct'].min(),
        'max_humidity': df['humidity_pct'].max(),
        'total_rain': df['rain_24h_mm'].sum(),
        'max_rain': df['rain_24h_mm'].max(),
        'rainy_days': (df['rain_24h_mm'] > 0).sum(),
        'data_completeness': (df['temperature_c'].notna().sum() / len(df)) * 100
    }
    
    # Write to Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # 1. Daily Data Sheet
        df_export = df[['date', 'day', 'temperature_c', 'humidity_pct', 'rain_24h_mm', 'battery_v', 'solar_v']]
        df_export.columns = ['Date', 'Day', 'Temperature (°C)', 'Humidity (%)', 'Rain 24hr (mm)', 'Battery (V)', 'Solar (V)']
        df_export.to_excel(writer, sheet_name='Daily Data', index=False)
        
        # 2. Summary Sheet
        summary_data = {
            'รายการ': [
                'รหัสสถานี', 'ชื่อสถานี', 'ปี', 'เดือน',
                '─────────────────────────',
                'อุณหภูมิเฉลี่ย (°C)', 'อุณหภูมิต่ำสุด (°C)', 'อุณหภูมิสูงสุด (°C)',
                '─────────────────────────',
                'ความชื้นเฉลี่ย (%)', 'ความชื้นต่ำสุด (%)', 'ความชื้นสูงสุด (%)',
                '─────────────────────────',
                'ปริมาณฝนรวม (mm)', 'ปริมาณฝนสูงสุด/วัน (mm)', 'จำนวนวันที่มีฝน',
                '─────────────────────────',
                'ความสมบูรณ์ของข้อมูล (%)'
            ],
            'ค่า': [
                station_info['code'],
                station_info['name'],
                year,
                f"{month} ({THAI_MONTHS[month]})",
                '',
                f"{stats['avg_temp']:.2f}" if pd.notna(stats['avg_temp']) else 'N/A',
                f"{stats['min_temp']:.2f}" if pd.notna(stats['min_temp']) else 'N/A',
                f"{stats['max_temp']:.2f}" if pd.notna(stats['max_temp']) else 'N/A',
                '',
                f"{stats['avg_humidity']:.2f}" if pd.notna(stats['avg_humidity']) else 'N/A',
                f"{stats['min_humidity']:.2f}" if pd.notna(stats['min_humidity']) else 'N/A',
                f"{stats['max_humidity']:.2f}" if pd.notna(stats['max_humidity']) else 'N/A',
                '',
                f"{stats['total_rain']:.2f}" if pd.notna(stats['total_rain']) else 'N/A',
                f"{stats['max_rain']:.2f}" if pd.notna(stats['max_rain']) else 'N/A',
                f"{stats['rainy_days']}" if pd.notna(stats['rainy_days']) else 'N/A',
                '',
                f"{stats['data_completeness']:.1f}%"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # 3. Monthly Comparison Sheet (placeholder for historical data)
        comparison_df = pd.DataFrame({
            'Month': [f"{year}-{month:02d}"],
            'Avg Temp (°C)': [stats['avg_temp']],
            'Total Rain (mm)': [stats['total_rain']],
            'Avg Humidity (%)': [stats['avg_humidity']]
        })
        comparison_df.to_excel(writer, sheet_name='Charts Data', index=False, startrow=1)
        
        # Get the workbook and sheets
        workbook = writer.book
        daily_sheet = writer.sheets['Daily Data']
        summary_sheet = writer.sheets['Summary']
        charts_sheet = writer.sheets['Charts Data']
        
        # Format Daily Data sheet
        for col in daily_sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            daily_sheet.column_dimensions[col_letter].width = adjusted_width
        
        # Add header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in daily_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format Summary sheet
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 25
        
        # Add formatting to summary headers
        for row in summary_sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
        
        # Create Temperature Chart
        if len(df) > 0 and df['temperature_c'].notna().any():
            temp_chart = LineChart()
            temp_chart.title = f"Daily Temperature - {THAI_MONTHS[month]} {year}"
            temp_chart.y_axis.title = "Temperature (°C)"
            temp_chart.x_axis.title = "Day"
            temp_chart.width = 15
            temp_chart.height = 8
            
            # Add data
            data_ref = Reference(daily_sheet, min_col=3, min_row=1, 
                               max_row=len(df)+1, max_col=3)
            cats_ref = Reference(daily_sheet, min_col=2, min_row=2, 
                                max_row=len(df)+1)
            temp_chart.add_data(data_ref, titles_from_data=True)
            temp_chart.set_categories(cats_ref)
            
            charts_sheet.add_chart(temp_chart, "E5")
        
        # Create Rainfall Chart
        if len(df) > 0 and df['rain_24h_mm'].notna().any():
            rain_chart = BarChart()
            rain_chart.title = f"Daily Rainfall - {THAI_MONTHS[month]} {year}"
            rain_chart.y_axis.title = "Rainfall (mm)"
            rain_chart.x_axis.title = "Day"
            rain_chart.width = 15
            rain_chart.height = 8
            
            # Add data
            data_ref = Reference(daily_sheet, min_col=5, min_row=1,
                               max_row=len(df)+1, max_col=5)
            cats_ref = Reference(daily_sheet, min_col=2, min_row=2,
                                max_row=len(df)+1)
            rain_chart.add_data(data_ref, titles_from_data=True)
            rain_chart.set_categories(cats_ref)
            
            charts_sheet.add_chart(rain_chart, "E20")
    
    print(f"✅ Excel file created: {filename}")
    return filename

def export_to_csv_simple(data: List[Dict], station_info: Dict,
                         year: int, month: int,
                         filename: str = None) -> str:
    """Export data to simple CSV file"""
    
    if filename is None:
        filename = f"weather_{station_info['code']}_{year}_{month:02d}.csv"
    
    # Prepare data for CSV
    csv_data = []
    for row in data:
        csv_data.append({
            'Station Code': station_info['code'],
            'Station Name': station_info['name'],
            'Year': year,
            'Month': month,
            'Date': row['date'],
            'Temperature (C)': row.get('temperature_c', ''),
            'Humidity (%)': row.get('humidity_pct', ''),
            'Rain 24hr (mm)': row.get('rain_24h_mm', ''),
            'Battery (V)': row.get('battery_v', ''),
            'Solar (V)': row.get('solar_v', '')
        })
    
    # Write CSV
    if csv_data:
        keys = csv_data[0].keys()
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(csv_data)
    
    print(f"✅ CSV file created: {filename}")
    return filename

# ---------------- Main Interactive Function ----------------
def main():
    """Main function with interactive menu"""
    
    print("=" * 70)
    print("🌧️  EEC Weather Data Monthly Extractor")
    print("    ระบบดึงข้อมูลสภาพอากาศรายเดือน")
    print("=" * 70)
    
    try:
        # Login
        print("\n📡 กำลังเชื่อมต่อระบบ EEC...")
        session = login("User", "User@1234", debug=False)
        print("✅ เชื่อมต่อสำเร็จ")
        
        # Fetch all stations
        print("\n🔍 กำลังดึงรายชื่อสถานีทั้งหมด...")
        stations = fetch_all_stations(session, debug=False)
        
        if not stations:
            print("❌ ไม่พบสถานีในระบบ!")
            return
        
        print(f"✅ พบสถานีทั้งหมด {len(stations)} สถานี")
        
        # Display stations in groups
        print("\n📍 รายชื่อสถานี:")
        print("─" * 70)
        
        station_list = list(stations.values())
        
        # Show stations in columns
        for i in range(0, min(30, len(station_list)), 3):
            row = ""
            for j in range(3):
                if i + j < len(station_list):
                    st = station_list[i + j]
                    row += f"[{st['code']:8s}] {st['name'][:20]:20s}  "
            print(row)
        
        if len(station_list) > 30:
            print(f"\n... และอีก {len(station_list) - 30} สถานี")
        
        # Station selection
        print("\n" + "=" * 70)
        while True:
            station_code = input("📌 ป้อนรหัสสถานี (เช่น EEC001): ").strip().upper()
            
            if station_code in stations:
                selected_station = stations[station_code]
                break
            else:
                print(f"❌ ไม่พบสถานี '{station_code}' กรุณาลองใหม่")
        
        print(f"✅ เลือกสถานี: [{selected_station['code']}] {selected_station['name']}")
        print(f"   สถานะล่าสุด: {selected_station.get('status', 'N/A')}")
        print(f"   อัพเดทล่าสุด: {selected_station.get('last_update', 'N/A')}")
        
        # Date selection
        print("\n📅 เลือกช่วงเวลา:")
        
        current_year = datetime.now().year
        year = int(input(f"   ปี (เช่น {current_year}): ") or current_year)
        
        print("\n   เดือน:")
        for i in range(1, 13, 3):
            row = ""
            for j in range(3):
                if i + j <= 12:
                    row += f"   {i+j:2d}. {THAI_MONTHS[i+j]:15s}"
            print(row)
        
        month = int(input("\n   เลือกเดือน (1-12): "))
        
        if month < 1 or month > 12:
            print("❌ เดือนไม่ถูกต้อง!")
            return
        
        # Calculate date range
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Fetch data
        print(f"\n⏳ กำลังดึงข้อมูล {THAI_MONTHS[month]} {year}")
        print(f"   ตั้งแต่วันที่ {start_date.date()} ถึง {end_date.date()}")
        print("   " + "─" * 50)
        
        daily_data = fetch_station_daily_data(session, selected_station['code'],
                                             start_date, end_date, debug=False)
        
        if not daily_data:
            print("❌ ไม่พบข้อมูลในช่วงเวลานี้!")
            return
        
        print(f"\n✅ ดึงข้อมูลสำเร็จ {len(daily_data)} วัน")
        
        # Display summary
        df = pd.DataFrame(daily_data)
        valid_temps = df['temperature_c'].dropna()
        valid_humidity = df['humidity_pct'].dropna()
        valid_rain = df['rain_24h_mm'].dropna()
        
        print("\n📊 สรุปข้อมูล:")
        print("─" * 50)
        
        if len(valid_temps) > 0:
            print(f"🌡️  อุณหภูมิ:")
            print(f"   - เฉลี่ย: {valid_temps.mean():.1f}°C")
            print(f"   - ต่ำสุด: {valid_temps.min():.1f}°C")
            print(f"   - สูงสุด: {valid_temps.max():.1f}°C")
        
        if len(valid_humidity) > 0:
            print(f"💧 ความชื้น:")
            print(f"   - เฉลี่ย: {valid_humidity.mean():.1f}%")
            print(f"   - ต่ำสุด: {valid_humidity.min():.1f}%")
            print(f"   - สูงสุด: {valid_humidity.max():.1f}%")
        
        if len(valid_rain) > 0:
            print(f"🌧️  ปริมาณฝน:")
            print(f"   - รวมทั้งเดือน: {valid_rain.sum():.1f} mm")
            print(f"   - สูงสุด/วัน: {valid_rain.max():.1f} mm")
            print(f"   - จำนวนวันที่มีฝน: {(valid_rain > 0).sum()} วัน")
        
        # Export options
        print("\n💾 ตัวเลือกการบันทึกข้อมูล:")
        print("   1. บันทึกเป็น Excel พร้อมกราฟ (.xlsx)")
        print("   2. บันทึกเป็น CSV (.csv)")
        print("   3. บันทึกทั้ง Excel และ CSV")
        print("   4. ไม่บันทึก")
        
        choice = input("\nเลือก (1-4): ").strip()
        
        output_dir = "weather_data"
        os.makedirs(output_dir, exist_ok=True)
        
        if choice in ['1', '3']:
            excel_file = os.path.join(output_dir, 
                                     f"weather_{selected_station['code']}_{year}_{month:02d}.xlsx")
            export_to_excel_with_charts(daily_data, selected_station, year, month, excel_file)
            print(f"   📁 ไฟล์: {excel_file}")
        
        if choice in ['2', '3']:
            csv_file = os.path.join(output_dir,
                                   f"weather_{selected_station['code']}_{year}_{month:02d}.csv")
            export_to_csv_simple(daily_data, selected_station, year, month, csv_file)
            print(f"   📁 ไฟล์: {csv_file}")
        
        print("\n" + "=" * 70)
        print("✨ ดำเนินการเสร็จสิ้น!")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n❌ เกิดข้อผิดพลาด: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
